#!/usr/bin/env python3
"""Babylon.js referring-expression data collection app.

Run:
    python3 study1_referring_expression.py

Then open http://127.0.0.1:5000. Responses are written to
task_graph/referring_expression_responses.csv and
task_graph/referring_expression_responses.xlsx.
"""

from __future__ import annotations

import csv
import base64
import binascii
import json
import re
import threading
import zipfile
from datetime import datetime
from pathlib import Path
from xml.sax.saxutils import escape, quoteattr

from flask import Flask, Response, abort, jsonify, request, send_from_directory


ROOT = Path(__file__).resolve().parent
ASSET_DIR = ROOT / "robot_assets" / "gearbox_parts"
PART_DIR = ASSET_DIR / "completed" / "colored_stl"
CSV_PATH = ROOT / "task_graph" / "referring_expression_responses.csv"
XLSX_PATH = ROOT / "task_graph" / "referring_expression_responses.xlsx"
RENDER_DIR = ROOT / "task_graph" / "referring_expression_renderings"
DETECTION_RENDER_DIR = ROOT / "task_graph" / "referring_expression_detection_renderings"
WRITE_LOCK = threading.Lock()
RENDER_LOCK = threading.Lock()
SELECTION_LOCK = threading.Lock()
GRAPH_SELECTION = {"revision": 0, "target": None}
STUDY_FILTER = {"revision": 0, "individual_parts_only": True}
REVIEW_SELECTION = {"revision": 0, "active": False, "response_index": None,
                    "target_id": None, "capture": False, "draw_boxes": False}
FIELDS = [
    "index", "timestamp", "participant", "target_type", "target_id", "target_name",
    "render_index", "description", "referring_type", "traits_used",
    "presentation_number", "target_presentations", "Verified",
]

# Rendering-class STL stem -> task-graph naming convention used in
# task_description.md, so evaluators never have to translate between two
# naming schemes. target_id (the .stl filename actually served/rendered)
# is unaffected; only the human/eval-facing target_name changes. Several
# STL assets are reused across multiple task-graph parts (see
# task_graph/part_naming_mapping.md), so BEARING/PIN/SCREW_ROW{n} are the
# generic roots task_description.md already uses for those groups
# (`BEARING_*`, `PIN_*`, `SCREW_ROW{n}_*`) rather than one specific
# row/side identifier.
TASKGRAPH_NAMES = {
    "BaseBoard": "BASE_BOARD",
    "Bearing": "BEARING",
    "Handle": "CRANK_HANDLE_ROW1",
    "Row1_GearRod": "GEAR_ROD_ROW1",
    "Row1_GearStand_Left": "STAND_ROW1_LEFT",
    "Row1_GearStand_Right": "STAND_ROW1_RIGHT",
    "Row1_Gear_Left": "GEAR_ROW1_LEFT",
    "Row1_Screws": "SCREW_ROW1",
    "Row2_GearRod": "GEAR_ROD_ROW2",
    "Row2_GearStand_Left": "STAND_ROW2_LEFT",
    "Row2_GearStand_Right": "STAND_ROW2_RIGHT",
    "Row2_Gear_Left": "GEAR_ROW2_LEFT",
    "Row2_Gear_Right": "GEAR_ROW2_RIGHT",
    "Row2_Screws": "SCREW_ROW2",
    "Row3_GearRod": "GEAR_ROD_ROW3",
    "Row3_GearStand_Left": "STAND_ROW3_LEFT",
    "Row3_GearStand_Right": "STAND_ROW3_RIGHT",
    "Row3_Gear_Left": "GEAR_ROW3_LEFT",
    "Row3_Gear_Right": "GEAR_ROW3_RIGHT",
    "Row3_Screws": "SCREW_ROW3",
    "Row4_GearRod": "GEAR_ROD_ROW4",
    "Row4_GearStand_Left": "STAND_ROW4_LEFT",
    "Row4_GearStand_Right": "STAND_ROW4_RIGHT",
    "Row4_Gear_Left": "GEAR_ROW4_LEFT",
    "Row4_Screws": "SCREW_ROW4",
    "WoodenPin": "PIN",
}

app = Flask(__name__)


def part_files() -> list[str]:
    return sorted(path.name for path in PART_DIR.glob("*.stl"))


def _semantic_obj_names() -> dict[str, str]:
    """Map task-graph raw-part IDs to mesh names in the colored assembly OBJ."""
    # Verified by geometry (vertex/face counts and bounds) against
    # GearBoxAssembly_Uncolored.obj. The two OBJ files do not retain identical
    # group order, so pairing their groups by position is unsafe.
    prefix_to_colored = {
        "BaseBoard": "Part_1.013",
        "Row1_Bearing_Left": "Part_1.024", "Row1_Bearing_Right": "Part_1.018",
        "Row1_GearStand_Left": "Part_1.004", "Row1_GearStand_Right": "Part_1.002",
        "Row1_GearRod": "Part_1.027", "Row1_Gear_Left": "Spur_gear_(42_teeth).001",
        "Row1_Pin_Left": "Part_1.017", "Row1_Pin_Right": "Part_1.011",
        "Row1_Screw_Left": "90751A122_18-8_Stainless_Steel_Socket_Head_Screws",
        "Row1_Screw_Right": "90751A122_18-8_Stainless_Steel_Socket_Head_Screws.001",
        "Row1_Handle": "Part_1.022",
        "Row2_Bearing_Left": "Part_1.023", "Row2_Bearing_Right": "Part_1.005",
        "Row2_GearStand_Left": "Part_1.026", "Row2_GearStand_Right": "Part_1.014",
        "Row2_GearRod": "Part_1.029", "Row2_Gear_Left": "Spur_gear_(20_teeth).001",
        "Row2_Gear_Right": "Spur_gear_(32_teeth)",
        "Row2_Pin_Left": "Part_1.012", "Row2_Pin_Right": "Part_1.006",
        "Row2_Screw_Left": "90358A118_Ultra-Low-Profile_Socket_Head_Screw",
        "Row2_Screw_Right": "90358A118_Ultra-Low-Profile_Socket_Head_Screw.001",
        "Row3_Bearing_Left": "Part_1", "Row3_Bearing_Right": "Part_1.016",
        "Row3_GearStand_Left": "Part_1.015", "Row3_GearStand_Right": "Part_1.003",
        "Row3_GearRod": "Part_1.028", "Row3_Gear_Left": "Spur_gear_(20_teeth)",
        "Row3_Gear_Right": "Spur_gear_(24_teeth)",
        "Row3_Pin_Left": "Part_1.010", "Row3_Pin_Right": "Part_1.001",
        "Row3_Screw_Left": "90942A121_High-Strength_A286_Stainless_Steel_Button_Head_Torx_Screw.001",
        "Row3_Screw_Right": "90942A121_High-Strength_A286_Stainless_Steel_Button_Head_Torx_Screw",
        "Row4_Bearing_Left": "Part_1.019", "Row4_Bearing_Right": "Part_1.008",
        "Row4_GearStand_Left": "Part_1.009", "Row4_GearStand_Right": "Part_1.007",
        "Row4_GearRod": "Part_1.021", "Row4_Gear_Left": "Spur_gear_(42_teeth)",
        "Row4_Pin_Left": "Part_1.020",
        "Row4_Screw_Left": "92000A107_Passivated_18-8_Stainless_Steel_Pan_Head_Phillips_Screws.001",
        "Row4_Screw_Right": "92000A107_Passivated_18-8_Stainless_Steel_Pan_Head_Phillips_Screws",
    }

    def prefix(part: str) -> str:
        if part == "BASE_BOARD":
            return "BaseBoard"
        tokens = part.split("_")
        if tokens[:2] == ["CRANK", "HANDLE"]:
            return f"Row{tokens[2][3:]}_Handle"
        kind = tokens[0]
        row = tokens[1][3:]
        side = f"_{tokens[2].title()}" if len(tokens) > 2 else ""
        label = {"STAND": "GearStand", "GEAR": "Gear", "GEARROD": "GearRod"}.get(kind, kind.title())
        return f"Row{row}_{label}{side}"

    from task_graph.gearbox_task_graph import PROVIDED_PARTS
    result = {}
    for part in PROVIDED_PARTS:
        key = part.replace("GEAR_ROD", "GEARROD")
        object_name = prefix_to_colored.get(prefix(key))
        if object_name is None:
            raise RuntimeError(f"No assembly mesh for task-graph part {part}")
        result[part] = object_name
    return result


def subassembly_targets() -> list[dict[str, object]]:
    """Build study targets directly from task-graph transformations."""
    from task_graph.gearbox_task_graph import build_steps

    raw_objects = _semantic_obj_names()
    constituents: dict[str, list[str]] = {
        part: [object_name] for part, object_name in raw_objects.items()
    }
    displayed: dict[str, list[str]] = {
        part: [object_name] for part, object_name in raw_objects.items()
    }
    targets = []
    for step in build_steps():
        meshes = []
        display_meshes = []
        for input_part in step.inputs:
            for mesh in constituents[input_part]:
                if mesh not in meshes:
                    meshes.append(mesh)
            for mesh in displayed[input_part]:
                if mesh not in display_meshes:
                    display_meshes.append(mesh)
        # Context objects are not consumed into the output, but are important
        # spatial references. For example, FASTENED_STAND and MOUNTED_ROW
        # targets should visibly include the baseboard they are mounted on.
        for context_part in step.context:
            for mesh in displayed[context_part]:
                if mesh not in display_meshes:
                    display_meshes.append(mesh)
        constituents[step.output] = meshes
        displayed[step.output] = display_meshes
        if step.output != "COMPLETED_GEARBOX_ASSEMBLY":
            targets.append({
                "file": step.output,
                "name": step.output,
                "kind": "subassembly",
                "assembly_objects": meshes,
                "display_objects": display_meshes,
            })
    return targets


def study_targets() -> list[dict[str, object]]:
    individuals = [{"file": name,
                     "name": TASKGRAPH_NAMES.get(Path(name).stem, Path(name).stem),
                     "kind": "part"}
                   for name in part_files()]
    targets = individuals + subassembly_targets()
    for render_index, target in enumerate(targets, start=1):
        target["render_index"] = render_index
    return targets


def clean_cell(value: object, limit: int = 5000) -> str:
    text = str(value or "").strip()[:limit]
    # Prevent spreadsheet formula injection in exported data.
    return "'" + text if text.startswith(("=", "+", "-", "@")) else text


def migrate_csv_schema() -> list[dict[str, str]]:
    """Return saved responses in the current schema, migrating older CSVs."""
    if not CSV_PATH.exists() or CSV_PATH.stat().st_size == 0:
        return []
    with CSV_PATH.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        source_fields = reader.fieldnames or []
        source_rows = list(reader)
    if source_fields == FIELDS:
        return source_rows

    migrated = []
    render_indices = {str(target["file"]): str(target["render_index"])
                      for target in study_targets()}
    for old in source_rows:
        # Old "skip" records had no participant description and are no longer
        # part of the response dataset.
        if old.get("action", "response") == "skip":
            continue
        old_timestamp = old.get("timestamp", old.get("timestamp_utc", ""))
        timestamp = old_timestamp[:19].replace("T", " ")
        target_id = old.get("target_id", old.get("part_file", ""))
        migrated.append({
            "index": str(len(migrated) + 1),
            "timestamp": timestamp,
            "participant": old.get("participant", ""),
            "target_type": old.get("target_type", "individual_part"),
            "target_id": target_id,
            "target_name": old.get("target_name", old.get("part_name", "")),
            "render_index": old.get("render_index", render_indices.get(target_id, "")),
            "description": old.get("description", ""),
            "referring_type": old.get("referring_type", ""),
            "traits_used": old.get("traits_used", ""),
            "presentation_number": old.get("presentation_number", ""),
            "target_presentations": old.get("target_presentations", ""),
            "Verified": old.get("Verified", ""),
        })
    with CSV_PATH.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(migrated)
    return migrated


def append_csv(row: dict[str, object]) -> None:
    new_file = not CSV_PATH.exists() or CSV_PATH.stat().st_size == 0
    with CSV_PATH.open("a", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS, lineterminator="\n")
        if new_file:
            writer.writeheader()
        writer.writerow(row)


def rebuild_xlsx() -> bool:
    with CSV_PATH.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    try:
        from openpyxl import Workbook
    except ImportError:
        return _rebuild_xlsx_stdlib(rows)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Responses"
    last_column = "M"
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = "Referring-expression classification scheme"
    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = ("Atomic traits: color | spatial | shape | relationship | size | "
                   "technical_name | function | material")
    sheet.merge_cells(f"A3:{last_column}3")
    sheet["A3"] = ("Use referring_type=mixed when more than one trait is used; "
                   "traits_used lists the participating traits separated by |")
    sheet.append(FIELDS)
    sheet.freeze_panes = "A5"
    for row in rows:
        sheet.append([row[field] for field in FIELDS])
    sheet.auto_filter.ref = f"A4:{last_column}{max(4, sheet.max_row)}"
    workbook.save(XLSX_PATH)
    return True


def _rebuild_xlsx_stdlib(rows: list[dict[str, str]]) -> bool:
    """Write the simple response workbook without requiring openpyxl."""
    scheme_rows = [
        "Referring-expression classification scheme",
        ("Atomic traits: color | spatial | shape | relationship | size | "
         "technical_name | function | material"),
        ("Use referring_type=mixed when more than one trait is used; "
         "traits_used lists the participating traits separated by |"),
    ]

    def column_name(number: int) -> str:
        result = ""
        while number:
            number, remainder = divmod(number - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def inline_cell(reference: str, value: object, style: int = 0) -> str:
        text = escape(str(value or ""))
        preserve = ' xml:space="preserve"' if text != text.strip() else ""
        return (f'<c r={quoteattr(reference)} t="inlineStr" s="{style}">'
                f'<is><t{preserve}>{text}</t></is></c>')

    worksheet_rows = []
    for row_number, text in enumerate(scheme_rows, start=1):
        worksheet_rows.append(
            f'<row r="{row_number}" ht="22" customHeight="1">'
            f'{inline_cell(f"A{row_number}", text, 1)}</row>'
        )
    header_cells = "".join(
        inline_cell(f"{column_name(column)}4", field, 2)
        for column, field in enumerate(FIELDS, start=1)
    )
    worksheet_rows.append(f'<row r="4">{header_cells}</row>')
    for row_number, row in enumerate(rows, start=5):
        cells = "".join(
            inline_cell(f"{column_name(column)}{row_number}", row.get(field, ""))
            for column, field in enumerate(FIELDS, start=1)
        )
        worksheet_rows.append(f'<row r="{row_number}">{cells}</row>')

    last_row = max(4, len(rows) + 4)
    worksheet = f'''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <sheetViews><sheetView workbookViewId="0"><pane ySplit="4" topLeftCell="A5" activePane="bottomLeft" state="frozen"/></sheetView></sheetViews>
  <cols><col min="1" max="13" width="18" customWidth="1"/><col min="8" max="8" width="60" customWidth="1"/><col min="10" max="10" width="34" customWidth="1"/></cols>
  <sheetData>{''.join(worksheet_rows)}</sheetData>
  <autoFilter ref="A4:M{last_row}"/>
  <mergeCells count="3"><mergeCell ref="A1:M1"/><mergeCell ref="A2:M2"/><mergeCell ref="A3:M3"/></mergeCells>
</worksheet>'''
    content_types = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>
  <Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>
  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>
</Types>'''
    package_rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>
</Relationships>'''
    workbook = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">
  <sheets><sheet name="Responses" sheetId="1" r:id="rId1"/></sheets>
</workbook>'''
    workbook_rels = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>
  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>
</Relationships>'''
    styles = '''<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="2"><font><sz val="11"/><name val="Calibri"/></font><font><b/><sz val="11"/><name val="Calibri"/></font></fonts>
  <fills count="2"><fill><patternFill patternType="none"/></fill><fill><patternFill patternType="gray125"/></fill></fills>
  <borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="3"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/><xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0"/><xf numFmtId="0" fontId="1" fillId="0" borderId="0" xfId="0" applyAlignment="1"><alignment horizontal="center"/></xf></cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
</styleSheet>'''
    try:
        with zipfile.ZipFile(XLSX_PATH, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("[Content_Types].xml", content_types)
            archive.writestr("_rels/.rels", package_rels)
            archive.writestr("xl/workbook.xml", workbook)
            archive.writestr("xl/_rels/workbook.xml.rels", workbook_rels)
            archive.writestr("xl/styles.xml", styles)
            archive.writestr("xl/worksheets/sheet1.xml", worksheet)
    except OSError:
        return False
    return True


@app.get("/")
def index() -> Response:
    return Response(INDEX_HTML, mimetype="text/html")


@app.get("/api/config")
def config() -> Response:
    targets = study_targets()
    return jsonify({
        "parts": targets,
        "assembly": "GearBoxAssembly.obj",
        "exploded": "GearBoxAssembly_Exploded.obj",
        "excel_enabled": excel_available(),
        "study_filter": dict(STUDY_FILTER),
    })


@app.get("/api/graph-selection")
def graph_selection() -> Response:
    with SELECTION_LOCK:
        return jsonify(dict(GRAPH_SELECTION))


@app.post("/api/graph-selection")
def set_graph_selection() -> Response:
    data = request.get_json(silent=True) or {}
    target = str(data.get("target", ""))
    valid = {str(item["file"]) for item in subassembly_targets()}
    if target not in valid:
        return jsonify({"error": "Unknown task-graph subassembly"}), 400
    with SELECTION_LOCK:
        GRAPH_SELECTION["revision"] += 1
        GRAPH_SELECTION["target"] = target
        return jsonify(dict(GRAPH_SELECTION))


@app.get("/api/study-filter")
def study_filter() -> Response:
    with SELECTION_LOCK:
        return jsonify(dict(STUDY_FILTER))


@app.post("/api/study-filter")
def set_study_filter() -> Response:
    data = request.get_json(silent=True) or {}
    individual_only = bool(data.get("individual_parts_only", False))
    with SELECTION_LOCK:
        if STUDY_FILTER["individual_parts_only"] != individual_only:
            STUDY_FILTER["revision"] += 1
            STUDY_FILTER["individual_parts_only"] = individual_only
        return jsonify(dict(STUDY_FILTER))


@app.get("/api/saved-responses")
def saved_responses() -> Response:
    with WRITE_LOCK:
        rows = migrate_csv_schema()
    return jsonify({"responses": rows})


@app.get("/api/review-selection")
def review_selection() -> Response:
    with SELECTION_LOCK:
        return jsonify(dict(REVIEW_SELECTION))


@app.get("/api/rendering-status")
def rendering_status() -> Response:
    try:
        response_index = int(request.args.get("response_index", ""))
    except ValueError:
        return jsonify({"error": "A valid response index is required"}), 400
    with WRITE_LOCK:
        exists = any(int(row["index"]) == response_index
                     for row in migrate_csv_schema())
    if not exists:
        return jsonify({"error": "Unknown response index"}), 404
    prefix = f"response_{response_index:06d}_"
    clean_images = sorted(path.name for path in RENDER_DIR.glob(prefix + "*.png"))
    detection_images = sorted(
        path.name for path in DETECTION_RENDER_DIR.glob(prefix + "*.png"))
    annotations = sorted(
        path.name for path in DETECTION_RENDER_DIR.glob(prefix + "*.json"))
    return jsonify({
        "response_index": response_index,
        "has_images": bool(clean_images or detection_images),
        "clean_images": clean_images,
        "detection_images": detection_images,
        "annotations": annotations,
    })


@app.post("/api/review-selection")
def set_review_selection() -> Response:
    data = request.get_json(silent=True) or {}
    active = bool(data.get("active", False))
    response_index = data.get("response_index")
    target_id = None
    if active:
        try:
            response_index = int(response_index)
        except (TypeError, ValueError):
            return jsonify({"error": "A valid response index is required"}), 400
        with WRITE_LOCK:
            row = next((item for item in migrate_csv_schema()
                        if int(item["index"]) == response_index), None)
        if row is None:
            return jsonify({"error": "Unknown response index"}), 404
        target_id = row["target_id"]
    with SELECTION_LOCK:
        REVIEW_SELECTION.update({
            "revision": REVIEW_SELECTION["revision"] + 1,
            "active": active,
            "response_index": response_index if active else None,
            "target_id": target_id,
            "capture": active and bool(data.get("capture", False)),
            "draw_boxes": active and bool(data.get("draw_boxes", False)),
        })
        return jsonify(dict(REVIEW_SELECTION))


@app.post("/api/renderings")
def save_rendering() -> Response:
    """Save a canvas-only PNG for the complete model or a study target."""
    data = request.get_json(silent=True) or {}
    target_id = str(data.get("target_id", ""))
    targets_by_id = {str(target["file"]): target for target in study_targets()}
    targets_by_id["COMPLETED_GEARBOX_ASSEMBLY"] = {
        "file": "COMPLETED_GEARBOX_ASSEMBLY", "render_index": 0,
    }
    if target_id != "complete_assembly" and target_id not in targets_by_id:
        return jsonify({"error": "Unknown rendering target"}), 400
    try:
        view = int(data.get("view", 1))
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid rendering view"}), 400
    if view not in (1, 2, 3):
        return jsonify({"error": "Rendering view must be 1, 2, or 3"}), 400
    response_index = data.get("response_index")
    if target_id != "complete_assembly":
        try:
            response_index = int(response_index)
        except (TypeError, ValueError):
            return jsonify({"error": "A response index is required"}), 400
        with WRITE_LOCK:
            rows = migrate_csv_schema()
        if not any(int(row["index"]) == response_index
                   and row["target_id"] == target_id for row in rows):
            return jsonify({"error": "Response index does not match this target"}), 400
    image = str(data.get("image", ""))
    prefix = "data:image/png;base64,"
    if not image.startswith(prefix):
        return jsonify({"error": "Expected a PNG data URL"}), 400
    try:
        png = base64.b64decode(image[len(prefix):], validate=True)
    except (binascii.Error, ValueError):
        return jsonify({"error": "Invalid PNG encoding"}), 400
    if not png.startswith(b"\x89PNG\r\n\x1a\n") or len(png) > 20_000_000:
        return jsonify({"error": "Invalid or oversized PNG"}), 400
    annotation = data.get("detection_annotation")
    detection_capture = isinstance(annotation, dict)
    normalized_annotation = None
    if detection_capture:
        try:
            width = int(annotation["image_width"])
            height = int(annotation["image_height"])
            boxes = annotation["boxes"]
            if width <= 0 or height <= 0 or not isinstance(boxes, list) or len(boxes) > 500:
                raise ValueError
            normalized_boxes = []
            for box in boxes:
                bbox = [round(float(value), 2) for value in box["bbox"]]
                if len(bbox) != 4 or bbox[2] < 0 or bbox[3] < 0:
                    raise ValueError
                normalized_boxes.append({
                    "label": str(box["label"])[:200],
                    "mesh_name": str(box["mesh_name"])[:200],
                    "bbox": bbox,
                })
            normalized_annotation = {
                "response_index": response_index,
                "target_id": target_id,
                "view": view,
                "image_width": width,
                "image_height": height,
                "bbox_format": "[x, y, width, height] in pixels from the top-left origin",
                "boxes": normalized_boxes,
            }
        except (KeyError, TypeError, ValueError):
            return jsonify({"error": "Invalid detection annotation"}), 400
    rendering_name = Path(target_id).stem if target_id.lower().endswith(".stl") else target_id
    with RENDER_LOCK:
        save_dir = DETECTION_RENDER_DIR if detection_capture else RENDER_DIR
        save_dir.mkdir(parents=True, exist_ok=True)
        if target_id == "complete_assembly":
            captures = []
            pattern = re.compile(r"^000_complete_assembly_capture(\d+)\.png$")
            for path in save_dir.glob("000_complete_assembly_capture*.png"):
                match = pattern.match(path.name)
                if match:
                    captures.append(int(match.group(1)))
            capture_index = max(captures, default=0) + 1
            filename = f"000_complete_assembly_capture{capture_index:03d}.png"
        else:
            filename = (f"response_{response_index:06d}_" +
                        f"target_{int(targets_by_id[target_id]['render_index']):03d}_" +
                        re.sub(r"[^A-Za-z0-9_.-]+", "_", rendering_name).strip("._") +
                        f"_view{view}.png")
        (save_dir / filename).write_bytes(png)
        annotation_filename = None
        if normalized_annotation is not None:
            annotation_filename = str(Path(filename).with_suffix(".json"))
            (save_dir / annotation_filename).write_text(
                json.dumps(normalized_annotation, indent=2) + "\n", encoding="utf-8")
    return jsonify({"ok": True, "filename": filename,
                    "annotation_filename": annotation_filename})


def excel_available() -> bool:
    # The standard-library fallback keeps Excel export available even when the
    # optional openpyxl package is not installed.
    return True


@app.get("/assets/<path:filename>")
def assets(filename: str) -> Response:
    return send_from_directory(ASSET_DIR, filename)


@app.get("/parts/<path:filename>")
def parts(filename: str) -> Response:
    if filename not in part_files():
        abort(404)
    return send_from_directory(PART_DIR, filename)


@app.post("/api/responses")
def save_response() -> Response:
    data = request.get_json(silent=True) or {}
    target_id = str(data.get("target_id", data.get("part_file", "")))
    action = str(data.get("action", "response"))
    targets_by_id = {str(target["file"]): target for target in study_targets()}
    if target_id not in targets_by_id:
        return jsonify({"error": "Unknown study target"}), 400
    target_info = targets_by_id[target_id]
    if action not in {"response", "skip"}:
        return jsonify({"error": "Invalid action"}), 400
    description = clean_cell(data.get("description"))
    if action == "response" and not description:
        return jsonify({"error": "Description is required"}), 400
    try:
        presentation = max(1, int(data.get("presentation_number", 1)))
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid presentation count"}), 400
    # Study 1 protocol is fixed at two presentations per part. Do not allow a
    # stale browser or edited request payload to create a different condition.
    target = 2
    if action == "skip":
        # Skipping only advances the study UI; it is not a response record.
        return jsonify({"ok": True, "skipped": True, "xlsx": excel_available()})
    with WRITE_LOCK:
        existing = migrate_csv_schema()
        index = max((int(item["index"]) for item in existing), default=0) + 1
        row = {
            "index": index,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "participant": clean_cell(data.get("participant"), 200),
            "target_type": ("subassembly" if target_info["kind"] == "subassembly"
                            else "individual_part"),
            "target_id": target_id,
            "target_name": str(target_info["name"]),
            "render_index": int(target_info["render_index"]),
            "description": description,
            "referring_type": "",
            "traits_used": "",
            "presentation_number": presentation,
            "target_presentations": target,
            "Verified": "",
        }
        append_csv(row)
        wrote_xlsx = rebuild_xlsx()
    return jsonify({"ok": True, "index": index, "xlsx": wrote_xlsx})


INDEX_HTML = r'''<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Gearbox Part Description Study</title>
  <script src="https://cdn.babylonjs.com/babylon.js"></script>
  <script src="https://cdn.babylonjs.com/loaders/babylonjs.loaders.min.js"></script>
  <style>
    :root { color-scheme: dark; --page:#101217; --panel:#171a20; --line:#343a46; --text:#f4f6fa; --field:#0e1117; --muted:#cbd2de; }
    body.light { color-scheme:light; --page:#fff; --panel:#f5f6f8; --line:#cbd0d8; --text:#17202d; --field:#fff; --muted:#485365; }
    * { box-sizing: border-box; }
    html, body { margin:0; width:100%; height:100%; overflow:hidden; font:14px/1.35 system-ui,sans-serif; background:var(--page); color:var(--text); }
    #app { display:grid; grid-template-columns:1fr 1fr; height:100%; }
    .viewer { position:relative; min-width:0; border-right:1px solid var(--line); }
    canvas { width:100%; height:100%; display:block; outline:none; touch-action:none; }
    .bar { position:absolute; z-index:2; top:12px; left:12px; right:12px; display:flex; gap:8px; align-items:center; flex-wrap:wrap; padding:9px; border:1px solid #ffffff24; border-radius:10px; background:#11151ddd; backdrop-filter:blur(8px); }
    .bar strong { margin-right:auto; }
    button, input, textarea { font:inherit; }
    button { color:var(--text); border:1px solid var(--line); border-radius:7px; background:var(--panel); padding:7px 11px; cursor:pointer; }
    button:hover { background:#384151; } button.primary { background:#2869db; } button.primary:hover { background:#3479ee; }
    button:disabled { opacity:.45; cursor:not-allowed; }
    #right { display:grid; grid-template-rows:minmax(0,1fr) auto; min-width:0; }
    #partView { position:relative; min-height:240px; }
    #form { padding:14px 16px 16px; background:var(--panel); border-top:1px solid var(--line); }
    .formrow { display:grid; grid-template-columns:minmax(160px,1fr) 150px; gap:12px; margin-bottom:10px; }
    label { display:block; color:var(--muted); font-size:12px; }
    input, textarea { width:100%; margin-top:4px; color:var(--text); background:var(--field); border:1px solid var(--line); border-radius:7px; padding:9px; }
    textarea { height:74px; resize:vertical; }
    .actions { display:flex; gap:8px; align-items:center; margin-top:9px; }
    #status { margin-left:auto; color:#abb5c5; min-height:20px; }
    #partLabel { color:#c9d8ff; overflow-wrap:anywhere; }
    #progress { color:#aeb8c8; }
    .hint { position:absolute; bottom:10px; left:12px; z-index:2; color:#d4dae4; background:#11151dcc; padding:6px 9px; border-radius:7px; pointer-events:none; }
    @media (max-width:850px) { #app { grid-template-columns:1fr; grid-template-rows:1fr 1fr; } #right { min-height:0; } .viewer { border-right:0; border-bottom:1px solid var(--line); } }
  </style>
</head>
<body>
<div id="app">
  <section class="viewer">
    <canvas id="assemblyCanvas"></canvas>
    <div class="bar"><strong>Full assembly</strong><button id="assembledBtn" class="primary">Assembled</button><button id="explodedBtn">Exploded</button><button id="saveCompleteView">Save complete view</button><button id="themeBtn">Light theme</button><button id="resetAssembly">Reset view</button></div>
    <div class="hint">The sampled part is boxed here. Click any component to inspect it.</div>
  </section>
  <section id="right">
    <div id="partView">
      <canvas id="partCanvas"></canvas>
      <div class="bar"><strong id="partLabel">Loading parts…</strong><span id="progress"></span><button id="resetPart">Reset view</button></div>
    </div>
    <div id="form">
      <div class="formrow">
        <label>Participant name<input id="participant" autocomplete="name" placeholder="Enter name"></label>
        <label>Descriptions per part<input id="target" type="number" value="2" readonly></label>
      </div>
      <label>Describe the highlighted part<textarea id="description" placeholder="Type the participant's description…"></textarea></label>
      <div class="actions"><button id="submit" class="primary">Save &amp; continue</button><button id="skip">Skip to next part</button><button id="newRound">Restart sampling</button><span id="status"></span></div>
    </div>
  </section>
</div>
<script>
(() => {
  const $ = id => document.getElementById(id);
  const PRESENTATIONS_PER_PART = 2;
  const FIXED_PRESENTATION_SEED = 20260826;
  const assemblyEngine = new BABYLON.Engine($('assemblyCanvas'), true, {preserveDrawingBuffer:true, stencil:true});
  const partEngine = new BABYLON.Engine($('partCanvas'), true, {preserveDrawingBuffer:true, stencil:true});
  let assemblyScene, assemblyCamera, partScene, currentPart, selectedAssemblyObject=null, allParts=[], parts=[], bag=[], counts={}, skipped={}, presentation=1, reviewActive=false;
  let activeAssemblyFile='GearBoxAssembly.obj'; const assemblyModels={}, assemblyLoads={};
  const capturedRenderings=new Set();
  const colors = {
    black:[.02,.02,.025], white:[.93,.93,.93], red:[.82,.035,.035], green:[.03,.55,.12],
    blue:[.035,.18,.86], brown:[.533,.322,.153], silver:[.53,.56,.60]
  };
  const assemblyParts = {
    'BaseBoard.stl':['Part_1.013'],
    'Bearing.stl':['Part_1','Part_1.005','Part_1.008','Part_1.016','Part_1.018','Part_1.019','Part_1.023','Part_1.024'],
    'Handle.stl':['Part_1.022'],
    'Row1_GearRod.stl':['Part_1.027'], 'Row1_GearStand_Left.stl':['Part_1.004'], 'Row1_GearStand_Right.stl':['Part_1.002'],
    'Row1_Gear_Left.stl':['Spur_gear_(42_teeth).001'], 'Row1_Screws.stl':['90751A122_18-8_Stainless_Steel_Socket_Head_Screws','90751A122_18-8_Stainless_Steel_Socket_Head_Screws.001'],
    'Row2_GearRod.stl':['Part_1.029'], 'Row2_GearStand_Left.stl':['Part_1.026'], 'Row2_GearStand_Right.stl':['Part_1.014'],
    'Row2_Gear_Left.stl':['Spur_gear_(20_teeth).001'], 'Row2_Gear_Right.stl':['Spur_gear_(32_teeth)'], 'Row2_Screws.stl':['90358A118_Ultra-Low-Profile_Socket_Head_Screw','90358A118_Ultra-Low-Profile_Socket_Head_Screw.001'],
    'Row3_GearRod.stl':['Part_1.028'], 'Row3_GearStand_Left.stl':['Part_1.015'], 'Row3_GearStand_Right.stl':['Part_1.003'],
    'Row3_Gear_Left.stl':['Spur_gear_(20_teeth)'], 'Row3_Gear_Right.stl':['Spur_gear_(24_teeth)'], 'Row3_Screws.stl':['90942A121_High-Strength_A286_Stainless_Steel_Button_Head_Torx_Screw','90942A121_High-Strength_A286_Stainless_Steel_Button_Head_Torx_Screw.001'],
    'Row4_Gear_Left.stl':['Spur_gear_(42_teeth)'], 'Row4_GearRod.stl':['Part_1.021'],
    'Row4_GearStand_Right.stl':['Part_1.007'], 'Row4_GearStand_Left.stl':['Part_1.009'],
    'Row4_Screws.stl':['92000A107_Passivated_18-8_Stainless_Steel_Pan_Head_Phillips_Screws','92000A107_Passivated_18-8_Stainless_Steel_Pan_Head_Phillips_Screws.001'],
    'WoodenPin.stl':['Part_1.001','Part_1.006','Part_1.010','Part_1.011','Part_1.012','Part_1.017','Part_1.020']
  };
  function sceneColor(){return document.body.classList.contains('light') ? new BABYLON.Color4(1,1,1,1) : new BABYLON.Color4(.055,.065,.085,1);}
  function setupScene(engine, canvas) {
    const scene = new BABYLON.Scene(engine); scene.clearColor = sceneColor();
    scene.ambientColor = new BABYLON.Color3(.1,.1,.1);
    scene.imageProcessingConfiguration.exposure = 1.0;
    scene.imageProcessingConfiguration.contrast = 1.12;
    const camera = new BABYLON.ArcRotateCamera('camera', -Math.PI/2, Math.PI/2.6, 4, BABYLON.Vector3.Zero(), scene);
    camera.attachControl(canvas, true); camera.wheelPrecision=45; camera.panningSensibility=1200; camera.lowerRadiusLimit=.01;
    const upper=new BABYLON.HemisphericLight('upperFill',new BABYLON.Vector3(0,1,0),scene); upper.intensity=.62; upper.groundColor=new BABYLON.Color3(.18,.2,.24);
    const lower=new BABYLON.HemisphericLight('lowerFill',new BABYLON.Vector3(0,-1,0),scene); lower.intensity=.08; lower.groundColor=new BABYLON.Color3(.1,.11,.13);
    const light = new BABYLON.DirectionalLight('key', new BABYLON.Vector3(-1,-2,-1), scene); light.intensity=.58;
    const cameraLight=new BABYLON.PointLight('cameraFill',BABYLON.Vector3.Zero(),scene); cameraLight.parent=camera; cameraLight.intensity=.24;
    return {scene,camera};
  }
  function makeAssemblyLabel(scene,text,anchorMesh,ownerRoot,color='#ffffff',width=.075,clearance=.018,screenLift=0,worldOffset=BABYLON.Vector3.Zero(),fixedPosition=null) {
    const plane=BABYLON.MeshBuilder.CreatePlane('spatial-label-'+text,{width,height:width*.30},scene);
    plane.billboardMode=BABYLON.Mesh.BILLBOARDMODE_ALL; plane.isPickable=false;
    plane.metadata={assemblySpatialLabel:true,anchorMesh,ownerRoot,clearance,screenLift,worldOffset,fixedPosition};
    const texture=new BABYLON.DynamicTexture('label-texture-'+text,{width:512,height:160},scene,true);
    texture.hasAlpha=true; texture.drawText(text,null,112,'bold 76px Arial',color,'rgba(8,12,20,.78)',true,true);
    const material=new BABYLON.StandardMaterial('label-material-'+text,scene);
    material.diffuseTexture=texture; material.opacityTexture=texture;
    material.emissiveColor=BABYLON.Color3.White(); material.disableLighting=true; material.backFaceCulling=false;
    plane.material=material; return plane;
  }
  function addBaseboardCoordinateFrame(scene,root,meshes) {
    const board=meshes.find(mesh=>mesh.name==='Part_1.013');
    root.metadata=root.metadata||{};
    if(!board||root.metadata.baseboardFrameAdded)return;
    root.metadata.baseboardFrameAdded=true;
    // Render the diagnostic frame after the assembly with a fresh depth buffer,
    // so axes lying in the board plane are not hidden inside its geometry.
    scene.setRenderingAutoClearDepthStencil(3,true,true,true);
    board.computeWorldMatrix(true);
    const bounds=board.getBoundingInfo().boundingBox;
    const origin=bounds.centerWorld.clone();
    const world=board.getWorldMatrix();
    const length=Math.max(.08,bounds.extendSizeWorld.length()*.28);
    const axes=[
      ['X',BABYLON.Axis.X,new BABYLON.Color3(1,.12,.12)],
      ['Y',BABYLON.Axis.Y,new BABYLON.Color3(.12,1,.2)],
      ['Z',BABYLON.Axis.Z,new BABYLON.Color3(.2,.45,1)],
    ];
    axes.forEach(([name,localAxis,color])=>{
      const direction=BABYLON.Vector3.TransformNormal(localAxis,world).normalize();
      const endpoint=origin.add(direction.scale(length));
      const line=BABYLON.MeshBuilder.CreateLines('baseboard-axis-'+name,{points:[origin,endpoint]},scene);
      line.color=color; line.isPickable=false;
      line.renderingGroupId=3;
      line.metadata={baseboardCoordinateFrame:true,ownerRoot:root};
      const plane=BABYLON.MeshBuilder.CreatePlane('baseboard-axis-label-'+name,{size:length*.22},scene);
      plane.position.copyFrom(endpoint.add(direction.scale(length*.12)));
      plane.billboardMode=BABYLON.Mesh.BILLBOARDMODE_ALL; plane.isPickable=false;
      plane.renderingGroupId=3;
      plane.metadata={baseboardCoordinateFrame:true,ownerRoot:root};
      const texture=new BABYLON.DynamicTexture('baseboard-axis-texture-'+name,{width:128,height:128},scene,true);
      texture.hasAlpha=true; texture.drawText(name,null,90,'bold 84px Arial',color.toHexString(),'transparent',true,true);
      const material=new BABYLON.StandardMaterial('baseboard-axis-material-'+name,scene);
      material.diffuseTexture=texture; material.opacityTexture=texture; material.emissiveColor=color;
      material.disableLighting=true; material.backFaceCulling=false; plane.material=material;
    });
  }
  function addAssemblySpatialLabels(scene,root,meshes) {
    if(root.metadata&&root.metadata.spatialLabelsAdded)return;
    root.metadata={...(root.metadata||{}),spatialLabelsAdded:true};
    const byName=name=>meshes.find(mesh=>mesh.name===name);
    const board=byName('Part_1.013');
    board.computeWorldMatrix(true);
    const boardY=BABYLON.Vector3.TransformNormal(BABYLON.Axis.Y,board.getWorldMatrix()).normalize();
    const boardOrigin=board.getBoundingInfo().boundingBox.centerWorld.clone();
    const leftPosition=boardOrigin.add(boardY.scale(.2));
    const rightPosition=boardOrigin.add(boardY.scale(-.2));
    const labels=[
      ['LEFT','Part_1.004','#ffe27a',.09,0,0,BABYLON.Vector3.Zero(),leftPosition], ['RIGHT','Part_1.002','#ffe27a',.09,0,0,BABYLON.Vector3.Zero(),rightPosition],
      ['ROW 1','Part_1.004','#ffffff',.068,.018,.045,BABYLON.Vector3.Zero(),null], ['ROW 2','Part_1.026','#ff7777',.068,.018,.045,BABYLON.Vector3.Zero(),null],
      ['ROW 3','Part_1.015','#62e58b',.068,.018,.045,BABYLON.Vector3.Zero(),null], ['ROW 4','Part_1.009','#6695ff',.068,.018,.045,BABYLON.Vector3.Zero(),null],
    ];
    labels.forEach(([text,meshName,color,width,clearance,screenLift,worldOffset,fixedPosition])=>{
      const anchor=byName(meshName); if(anchor)makeAssemblyLabel(scene,text,anchor,root,color,width,clearance,screenLift,worldOffset,fixedPosition);
    });
    if(!scene.metadata||!scene.metadata.spatialLabelObserver) {
      scene.metadata={...(scene.metadata||{}),spatialLabelObserver:true};
      scene.onBeforeRenderObservable.add(()=>{
        const suppressed=scene.metadata&&scene.metadata.suppressSpatialLabels;
        scene.meshes.filter(mesh=>mesh.metadata&&mesh.metadata.baseboardCoordinateFrame).forEach(mesh=>{
          mesh.setEnabled(!suppressed&&mesh.metadata.ownerRoot.isEnabled());
        });
        scene.meshes.filter(mesh=>mesh.metadata&&mesh.metadata.assemblySpatialLabel).forEach(label=>{
          const {anchorMesh,ownerRoot,clearance,screenLift,worldOffset,fixedPosition}=label.metadata;
          label.setEnabled(!suppressed&&ownerRoot.isEnabled());
          if(suppressed||!ownerRoot.isEnabled())return;
          if(fixedPosition){label.position.copyFrom(fixedPosition);return;}
          anchorMesh.computeWorldMatrix(true);
          const bounds=anchorMesh.getBoundingInfo().boundingSphere;
          const center=bounds.centerWorld;
          const towardCamera=scene.activeCamera.position.subtract(center).normalize();
          const cameraUp=scene.activeCamera.upVector.clone().normalize().scale(screenLift);
          label.position.copyFrom(center.add(towardCamera.scale(bounds.radiusWorld+clearance)).add(cameraUp).add(worldOffset));
        });
      });
    }
  }
  function frame(scene, meshes, camera=scene.activeCamera) {
    const visible=meshes.filter(m=>m.getTotalVertices && m.getTotalVertices()>0 && m.isEnabled());
    if(!visible.length) return;
    let min=new BABYLON.Vector3(Infinity,Infinity,Infinity), max=new BABYLON.Vector3(-Infinity,-Infinity,-Infinity);
    visible.forEach(m=>{ m.computeWorldMatrix(true); const b=m.getBoundingInfo().boundingBox; min=BABYLON.Vector3.Minimize(min,b.minimumWorld); max=BABYLON.Vector3.Maximize(max,b.maximumWorld); });
    const center=min.add(max).scale(.5), size=max.subtract(min).length(); camera.setTarget(center); camera.radius=Math.max(size*.72,.02); camera.minZ=Math.max(size/10000,.0001); camera.maxZ=Math.max(size*20,100);
  }
  function showBoxes(meshes, scene) {
    scene.meshes.forEach(m=>{if(m.showBoundingBox)m.showBoundingBox=false;});
    scene.forceShowBoundingBoxes=false;
    meshes.forEach(mesh=>mesh.showBoundingBox=true);
    const renderer=scene.getBoundingBoxRenderer();
    const light=document.body.classList.contains('light');
    renderer.frontColor=light ? new BABYLON.Color3(.02,.12,.65) : new BABYLON.Color3(1,.82,.05);
    renderer.backColor=renderer.frontColor;
  }
  function drawProjectedBoxes(context,canvas,scene,target) {
    const camera=scene.activeCamera;
    const viewport=camera.viewport.toGlobal(canvas.width,canvas.height);
    const transform=scene.getTransformMatrix();
    const meshes=scene.meshes.filter(mesh=>mesh.isEnabled()&&mesh.isVisible&&
      mesh.getTotalVertices&&mesh.getTotalVertices()>0&&
      !(mesh.metadata&&(mesh.metadata.assemblySpatialLabel||mesh.metadata.baseboardCoordinateFrame)));
    const annotations=[];
    context.save(); context.font='600 16px system-ui,sans-serif'; context.lineWidth=3;
    meshes.forEach((mesh,index)=>{
      mesh.computeWorldMatrix(true);
      const points=mesh.getBoundingInfo().boundingBox.vectorsWorld.map(point=>
        BABYLON.Vector3.Project(point,BABYLON.Matrix.Identity(),transform,viewport));
      const visible=points.filter(point=>point.z>=0&&point.z<=1);
      if(!visible.length)return;
      const xs=points.map(point=>point.x), ys=points.map(point=>point.y);
      const left=Math.max(0,Math.min(...xs)), top=Math.max(0,Math.min(...ys));
      const right=Math.min(canvas.width,Math.max(...xs)), bottom=Math.min(canvas.height,Math.max(...ys));
      if(right-left<2||bottom-top<2)return;
      // A projected 3D box can exist even when another part completely hides
      // the mesh. Sample picking rays across the rectangle and retain the box
      // only when this mesh is the nearest visible surface for at least one ray.
      const fractions=[.08,.25,.5,.75,.92];
      let visiblyHit=false;
      for(const fy of fractions){
        for(const fx of fractions){
          const pick=scene.pick(left+(right-left)*fx,top+(bottom-top)*fy,
            candidate=>candidate.isEnabled()&&candidate.isVisible&&
              candidate.getTotalVertices&&candidate.getTotalVertices()>0,false,camera);
          if(pick&&pick.hit&&pick.pickedMesh&&
              (pick.pickedMesh===mesh||pick.pickedMesh.isDescendantOf(mesh))){
            visiblyHit=true; break;
          }
        }
        if(visiblyHit)break;
      }
      if(!visiblyHit)return;
      const hue=(index*67)%360, color=`hsl(${hue} 95% 62%)`;
      const mapped=Object.entries(assemblyParts).find(([_file,names])=>names.includes(mesh.name));
      const label=target.kind==='part' ? target.name : (mapped?mapped[0].replace(/\.stl$/,''):mesh.name);
      annotations.push({label,mesh_name:mesh.name,bbox:[left,top,right-left,bottom-top]});
      context.strokeStyle=color; context.strokeRect(left,top,right-left,bottom-top);
      const labelWidth=context.measureText(label).width+10;
      const labelTop=Math.max(0,top-22);
      context.fillStyle='rgba(0,0,0,.78)'; context.fillRect(left,labelTop,labelWidth,22);
      context.fillStyle=color; context.fillText(label,left+5,labelTop+17);
    });
    context.restore();
    return {image_width:canvas.width,image_height:canvas.height,boxes:annotations};
  }
  async function captureRendering(targetId, renderIndex, canvas, scene, responseIndex=null, force=false, drawBoxes=false, target=null) {
    const captureKey=responseIndex===null ? targetId : targetId+'@'+responseIndex;
    if(capturedRenderings.has(captureKey)&&!force)return true;
    if(!canvas||!scene)return false;
    capturedRenderings.add(captureKey);
    const camera=scene.activeCamera, baseAlpha=camera.alpha, baseBeta=camera.beta, baseRadius=camera.radius;
    try {
      const savedFiles=[];
      const viewCount=1;
      for(let view=1;view<=viewCount;view++) {
        // Capture exactly one image from the participant's current camera pose.
        camera.alpha=baseAlpha; camera.beta=baseBeta; camera.radius=baseRadius;
        await new Promise(resolve=>requestAnimationFrame(resolve)); scene.render();
        let image, detectionAnnotation=null;
        if(responseIndex!==null) {
          const stamped=document.createElement('canvas'); stamped.width=canvas.width; stamped.height=canvas.height;
          const context=stamped.getContext('2d'); context.drawImage(canvas,0,0);
          if(drawBoxes&&target)detectionAnnotation=drawProjectedBoxes(context,stamped,scene,target);
          const fontSize=Math.max(20,Math.round(Math.min(canvas.width,canvas.height)*.045));
          const label='Response #'+String(responseIndex).padStart(6,'0')+' · Target #'+String(renderIndex).padStart(3,'0')+' · View 1/1'; context.font=`600 ${fontSize}px system-ui,sans-serif`;
          const width=context.measureText(label).width; const pad=Math.round(fontSize*.42);
          context.fillStyle='rgba(0,0,0,.72)'; context.fillRect(pad,pad,width+pad*2,fontSize+pad*1.35);
          context.fillStyle='white'; context.textBaseline='top'; context.fillText(label,pad*2,pad*1.45);
          image=stamped.toDataURL('image/png');
        } else image=canvas.toDataURL('image/png');
        const response=await fetch('/api/renderings',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({target_id:targetId,response_index:responseIndex,view,image,detection_annotation:detectionAnnotation})});
        const result=await response.json();
        if(!response.ok)throw new Error(result.error||'Rendering save failed');
        savedFiles.push(result.filename);
      }
      return savedFiles;
    } catch(error) {
      capturedRenderings.delete(captureKey);
      console.warn('Could not save rendering',targetId,error);
      return false;
    } finally {
      camera.alpha=baseAlpha; camera.beta=baseBeta; camera.radius=baseRadius; scene.render();
    }
  }
  async function saveCleanCompleteView() {
    if(activeAssemblyFile!=='GearBoxAssembly.obj')return false;
    if(!assemblyScene||!assemblyModels[activeAssemblyFile])return false;
    const boxed=assemblyScene.meshes.filter(mesh=>mesh.showBoundingBox);
    // Bounding boxes are renderer overlays, so disable both per-mesh flags and
    // the scene-wide force flag for the capture, then restore them afterward.
    const forced=assemblyScene.forceShowBoundingBoxes;
    const labels=assemblyScene.meshes.filter(mesh=>mesh.metadata&&(mesh.metadata.assemblySpatialLabel||mesh.metadata.baseboardCoordinateFrame)&&mesh.isEnabled());
    assemblyScene.metadata={...(assemblyScene.metadata||{}),suppressSpatialLabels:true};
    assemblyScene.meshes.forEach(mesh=>{mesh.showBoundingBox=false;});
    labels.forEach(mesh=>mesh.setEnabled(false));
    assemblyScene.forceShowBoundingBoxes=false;
    try {
      return await captureRendering('complete_assembly',0,$('assemblyCanvas'),assemblyScene,null,true);
    } finally {
      boxed.forEach(mesh=>{mesh.showBoundingBox=true;});
      labels.forEach(mesh=>mesh.setEnabled(true));
      assemblyScene.metadata.suppressSpatialLabels=false;
      assemblyScene.forceShowBoundingBoxes=forced;
      assemblyScene.render();
    }
  }
  function highlightSampledPart() {
    if(!assemblyScene||!currentPart)return;
    const model=assemblyModels[activeAssemblyFile];
    const mapped=currentPart.kind==='subassembly' ? currentPart.display_objects : (assemblyParts[currentPart.file]||[]);
    const wanted=selectedAssemblyObject&&mapped.includes(selectedAssemblyObject) ? selectedAssemblyObject : mapped[0];
    const names=currentPart.kind==='subassembly' ? mapped : [wanted];
    showBoxes(model ? model.meshes.filter(m=>names.includes(m.name)) : [],assemblyScene);
  }
  function partForAssemblyObject(name) {
    const entry=Object.entries(assemblyParts).find(([,names])=>names.includes(name));
    return entry ? parts.find(part=>part.kind==='part'&&part.file===entry[0]) : null;
  }
  function materialKey(file) {
    if(file==='BaseBoard.stl'||file==='Bearing.stl'||file==='Row2_Screws.stl'||file==='Row3_Screws.stl') return 'black';
    if(file==='Handle.stl'||file.startsWith('Row1_')&&!file.includes('Screws')) return 'white';
    if(file==='Row1_Screws.stl'||file==='Row4_Screws.stl') return 'silver';
    if(file==='WoodenPin.stl') return 'brown'; if(file.startsWith('Row2')) return 'red';
    if(file.startsWith('Row3')) return 'green'; if(file.startsWith('Row4')) return 'blue'; return 'white';
  }
  function applyPartMaterial(scene, meshes, file) {
    const key=materialKey(file), c=colors[key], mat=new BABYLON.PBRMaterial('part-'+key,scene);
    mat.albedoColor=new BABYLON.Color3(...c); mat.metallic=key==='silver' ? .9 : 0; mat.roughness=key==='silver' ? .28 : .55;
    meshes.forEach(m=>{if(m.getTotalVertices()>0)m.material=mat;});
  }
  async function loadAssembly(file, activate=true) {
    if(!assemblyScene) {
      const made=setupScene(assemblyEngine,$('assemblyCanvas')); assemblyScene=made.scene; assemblyCamera=made.camera;
      assemblyScene.onPointerObservable.add(info=>{
        if(info.type!==BABYLON.PointerEventTypes.POINTERPICK) return;
        const pick=info.pickInfo;
        if(pick && pick.hit && pick.pickedMesh && pick.pickedMesh.getTotalVertices()>0) {
          const part=partForAssemblyObject(pick.pickedMesh.name);
          if(part&&currentPart&&part.file===currentPart.file) {
            highlightSampledPart();
          } else if(part) {
            showError('Study order is fixed; finish the currently assigned part first.');
          } else showError('This assembly object has no matching individual study part.');
        }
      });
    }
    if(activate) {
      activeAssemblyFile=file;
      Object.values(assemblyModels).forEach(model=>model.root.setEnabled(false));
    }
    if(!assemblyModels[file]) {
      if(activate)$('status').textContent='Loading '+(file.includes('Exploded')?'exploded':'assembled')+' model once…';
      if(!assemblyLoads[file]) assemblyLoads[file]=(async()=>{
        const result=await BABYLON.SceneLoader.ImportMeshAsync('', '/assets/', file, assemblyScene);
        const root=new BABYLON.TransformNode('cached-'+file,assemblyScene);
        result.meshes.forEach(m=>{if(!m.parent)m.parent=root;});
        [...new Set(result.meshes.map(m=>m.material).filter(Boolean))].forEach(mat=>{
          if(mat.specularColor)mat.specularColor=new BABYLON.Color3(.12,.12,.12);
          if('specularPower' in mat)mat.specularPower=64;
        });
        assemblyModels[file]={root,meshes:result.meshes};
        if(file==='GearBoxAssembly.obj'){
          addBaseboardCoordinateFrame(assemblyScene,root,result.meshes);
          addAssemblySpatialLabels(assemblyScene,root,result.meshes);
        }
        if(activeAssemblyFile!==file)root.setEnabled(false);
      })();
      await assemblyLoads[file];
      if(activate)$('status').textContent='Model cached';
    }
    if(!activate)return;
    const model=assemblyModels[file];
    if(activeAssemblyFile!==file)return;
    model.root.setEnabled(true);
    frame(assemblyScene,model.meshes,assemblyCamera);
    highlightSampledPart();
  }
  async function loadPart(part, assemblyObject=null) {
    currentPart=part; selectedAssemblyObject=assemblyObject; presentation=(counts[part.file]||0)+1;
    $('submit').disabled=false; $('skip').disabled=false;
    highlightSampledPart();
    $('partLabel').textContent=part.name; updateProgress();
    if(partScene) partScene.dispose();
    const made=setupScene(partEngine,$('partCanvas')); partScene=made.scene;
    let result;
    if(part.kind==='assembly') {
      result=await BABYLON.SceneLoader.ImportMeshAsync('', '/assets/', 'GearBoxAssembly.obj', partScene);
    } else if(part.kind==='subassembly') {
      result=await BABYLON.SceneLoader.ImportMeshAsync(part.display_objects, '/assets/', 'GearBoxAssembly.obj', partScene);
    } else {
      result=await BABYLON.SceneLoader.ImportMeshAsync('', '/parts/', encodeURIComponent(part.file), partScene);
      applyPartMaterial(partScene,result.meshes,part.file);
    }
    frame(partScene,result.meshes,made.camera);
    highlightSampledPart();
    $('description').value=''; $('description').focus();
  }
  function fixedShuffle(items) {
    const result=items.slice(); let seed=FIXED_PRESENTATION_SEED>>>0;
    for(let i=result.length-1;i>0;i--){
      seed=(Math.imul(1664525,seed)+1013904223)>>>0;
      const j=seed%(i+1); [result[i],result[j]]=[result[j],result[i]];
    }
    return result;
  }
  function refillBag() {
    const target=PRESENTATIONS_PER_PART;
    bag=fixedShuffle(parts.filter(
      p=>(counts[p.file]||0)<target&&!skipped[p.file]));
  }
  function applyStudyFilter(individualOnly) {
    parts=individualOnly ? allParts.filter(part=>part.kind==='part') : allParts.slice();
    bag=[];
    if(currentPart&&!parts.some(part=>part.file===currentPart.file)){
      currentPart=null; selectedAssemblyObject=null; nextPart(true);
    } else refillBag();
    $('status').textContent=individualOnly?'Study filter: individual parts only':'Study filter: parts and subassemblies';
  }
  function nextPart(forceNew=false) {
    if(reviewActive)return;
    if(!bag.length) refillBag();
    if(!bag.length){ currentPart=null; $('partLabel').textContent='Sampling complete'; $('progress').textContent=''; $('submit').disabled=true; $('skip').disabled=true; return; }
    $('submit').disabled=false; $('skip').disabled=false; loadPart(bag.pop()).catch(showError);
  }
  function updateProgress(){const target=PRESENTATIONS_PER_PART, completed=Object.values(counts).reduce((a,b)=>a+b,0); $('progress').textContent=`showing ${presentation}/${target} · ${completed}/${parts.length*target} saved`;}
  async function applyReviewSelection(review) {
    reviewActive=Boolean(review.active);
    if(!reviewActive){
      $('submit').disabled=false; $('skip').disabled=false;
      $('status').textContent='CSV review mode closed'; nextPart(true); return;
    }
    let target=allParts.find(part=>part.file===review.target_id);
    if(review.target_id==='COMPLETED_GEARBOX_ASSEMBLY'){
      target={file:'COMPLETED_GEARBOX_ASSEMBLY',name:'COMPLETED_GEARBOX_ASSEMBLY',kind:'assembly',render_index:0};
    }
    if(!target){showError('Saved response target is unavailable: '+review.target_id);return;}
    // A capture command for the target already on screen must reuse the live
    // scene. Reloading it would call frame(), resetting the participant's
    // carefully chosen camera pose immediately before capture.
    const reuseCurrentView=Boolean(review.capture&&partScene&&currentPart&&
      currentPart.file===target.file);
    if(!reuseCurrentView)await loadPart(target);
    $('submit').disabled=true; $('skip').disabled=true;
    $('status').textContent='Reviewing response #'+review.response_index+' · '+target.name;
    if(review.capture){
      const saved=await captureRendering(target.file,target.render_index,$('partCanvas'),partScene,review.response_index,true,Boolean(review.draw_boxes),target);
      $('status').textContent=Array.isArray(saved)?('Captured response #'+review.response_index):('Capture failed for response #'+review.response_index);
    }
  }
  async function record(action) {
    if(!currentPart) return;
    const description=$('description').value.trim(); if(action==='response'&&!description){showError('Please enter a description, or use Skip.');return;}
    setBusy(true); const payload={participant:$('participant').value,target_id:currentPart.file,description,presentation_number:presentation,target_presentations:PRESENTATIONS_PER_PART,action};
    try { const res=await fetch('/api/responses',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)}); const data=await res.json(); if(!res.ok)throw new Error(data.error||'Save failed'); let captured=true;if(action==='response'){captured=await captureRendering(currentPart.file,currentPart.render_index,$('partCanvas'),partScene,data.index);counts[currentPart.file]=(counts[currentPart.file]||0)+1;}else skipped[currentPart.file]=true; $('status').textContent=action==='skip'?'Part skipped':(captured?'Response and views saved':'Response saved; image capture failed'); nextPart(action==='skip'); }
    catch(e){showError(e);} finally{setBusy(false);}
  }
  function setBusy(v){$('submit').disabled=v;$('skip').disabled=v;}
  function showError(e){$('status').textContent=e.message||String(e);$('status').style.color='#ff8d8d';setTimeout(()=>{$('status').style.color='';},3500);}
  async function init(){
    const cfg=await fetch('/api/config').then(r=>r.json()); allParts=cfg.parts;
    applyStudyFilter(Boolean(cfg.study_filter&&cfg.study_filter.individual_parts_only));
    if(!parts.length)throw new Error('No individual-part or subassembly study targets found');
    if(!cfg.excel_enabled)$('status').textContent='Saving CSV (install openpyxl for Excel too)';
    await loadAssembly(cfg.assembly);
    refillBag(); nextPart();
    loadAssembly(cfg.exploded,false).catch(showError);
    let graphRevision=-1, filterRevision=cfg.study_filter?cfg.study_filter.revision:-1, reviewRevision=-1;
    setInterval(async()=>{
      try {
        const [selection,filter,review]=await Promise.all([
          fetch('/api/graph-selection',{cache:'no-store'}).then(r=>r.json()),
          fetch('/api/study-filter',{cache:'no-store'}).then(r=>r.json()),
          fetch('/api/review-selection',{cache:'no-store'}).then(r=>r.json()),
        ]);
        if(filter.revision!==filterRevision){
          filterRevision=filter.revision;
          applyStudyFilter(Boolean(filter.individual_parts_only));
        }
        if(review.revision!==reviewRevision){
          reviewRevision=review.revision;
          await applyReviewSelection(review);
        }
        if(selection.revision===graphRevision)return;
        graphRevision=selection.revision;
        if(!selection.target||filter.individual_parts_only||review.active)return;
        const target=parts.find(part=>part.kind==='subassembly'&&part.file===selection.target);
        if(target){bag=bag.filter(item=>item.file!==target.file);await loadPart(target);$('status').textContent='Task graph selected '+target.name;}
      } catch(_e) { /* Flask may be restarting; retry on the next poll. */ }
    },250);
  }
  $('assembledBtn').onclick=()=>{ $('assembledBtn').classList.add('primary');$('explodedBtn').classList.remove('primary');loadAssembly('GearBoxAssembly.obj').catch(showError); };
  $('explodedBtn').onclick=()=>{ $('explodedBtn').classList.add('primary');$('assembledBtn').classList.remove('primary');loadAssembly('GearBoxAssembly_Exploded.obj').catch(showError); };
  $('resetAssembly').onclick=()=>{const model=assemblyModels[activeAssemblyFile];if(model)frame(assemblyScene,model.meshes);};
  $('saveCompleteView').onclick=async()=>{
    if(activeAssemblyFile!=='GearBoxAssembly.obj'){$('status').textContent='Switch to Assembled before saving the complete view';return;}
    const button=$('saveCompleteView'); button.disabled=true; button.textContent='Saving…';
    const saved=await saveCleanCompleteView();
    button.disabled=false; button.textContent='Save complete view';
    $('status').textContent=Array.isArray(saved)?('Saved '+saved[0]):'Complete-view capture failed';
  };
  $('resetPart').onclick=()=>partScene&&frame(partScene,partScene.meshes);
  $('themeBtn').onclick=()=>{const light=document.body.classList.toggle('light');$('themeBtn').textContent=light?'Dark theme':'Light theme';if(assemblyScene){assemblyScene.clearColor=sceneColor();highlightSampledPart();}if(partScene)partScene.clearColor=sceneColor();};
  $('submit').onclick=()=>record('response'); $('skip').onclick=()=>record('skip');
  $('newRound').onclick=()=>{counts={};skipped={};currentPart=null;refillBag();nextPart(true);$('status').textContent='Sampling restarted';};
  $('description').addEventListener('keydown',e=>{if((e.ctrlKey||e.metaKey)&&e.key==='Enter')record('response');});
  window.addEventListener('resize',()=>{assemblyEngine.resize();partEngine.resize();});
  assemblyEngine.runRenderLoop(()=>{if(assemblyScene)assemblyScene.render();}); partEngine.runRenderLoop(()=>{if(partScene)partScene.render();}); init().catch(showError);
})();
</script>
</body>
</html>'''


def main() -> None:
    if not PART_DIR.is_dir():
        raise SystemExit(f"Missing colored STL directory: {PART_DIR}")
    print("Open http://127.0.0.1:5000")
    print(f"Responses: {CSV_PATH}")
    if not excel_available():
        print("Tip: install openpyxl to also save an .xlsx workbook")
    app.run(host="127.0.0.1", port=5000, debug=False, threaded=True)


if __name__ == "__main__":
    main()
