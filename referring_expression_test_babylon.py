#!/usr/bin/env python3
"""Babylon.js referring-expression data collection app.

Run:
    python referring_expression_test_babylon.py

Then open http://127.0.0.1:5050. Responses are written to
referring_expression_responses.csv and, when openpyxl is installed, to
referring_expression_responses.xlsx as well.
"""

from __future__ import annotations

import argparse
import base64
import binascii
import csv
import json
import threading
import uuid
from datetime import datetime, timezone
from pathlib import Path

from flask import Flask, Response, abort, jsonify, request, send_from_directory

from referring_expression_resolver import Candidate, ReferringExpressionResolver, create_resolver


ROOT = Path(__file__).resolve().parent
ASSET_DIR = ROOT / "robot_assets" / "gearbox_parts"
PART_DIR = ASSET_DIR / "completed" / "colored_stl"
CSV_PATH = ROOT / "referring_expression_responses.csv"
XLSX_PATH = ROOT / "referring_expression_responses.xlsx"
PREDICTIONS_CSV_PATH = ROOT / "referring_expression_predictions.csv"
WRITE_LOCK = threading.Lock()
RESOLVER: ReferringExpressionResolver | None = None
FIELDS = [
    "response_id", "timestamp_utc", "participant", "part_file", "part_name",
    "description", "presentation_number", "target_presentations", "action",
]
PREDICTION_FIELDS = [
    "response_id", "timestamp_utc", "participant", "expression",
    "ground_truth_part_file", "predicted_part_file", "predicted_mesh_name",
    "is_correct", "confidence", "mapping_score", "latency_ms", "resolver", "model", "raw_output",
]

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024


def part_files() -> list[str]:
    return sorted(path.name for path in PART_DIR.glob("*.stl"))


def clean_cell(value: object, limit: int = 5000) -> str:
    text = str(value or "").strip()[:limit]
    # Prevent spreadsheet formula injection in exported data.
    return "'" + text if text.startswith(("=", "+", "-", "@")) else text


def optional_finite_number(value: object) -> float | None:
    if value is None or value == "":
        return None
    number = float(value)
    if number != number or number in {float("inf"), float("-inf")}:
        raise ValueError("Prediction metrics must be finite")
    return number


def append_csv(row: dict[str, object]) -> None:
    new_file = not CSV_PATH.exists() or CSV_PATH.stat().st_size == 0
    with CSV_PATH.open("a", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        if new_file:
            writer.writeheader()
        writer.writerow(row)


def append_prediction_csv(row: dict[str, object]) -> None:
    new_file = not PREDICTIONS_CSV_PATH.exists() or PREDICTIONS_CSV_PATH.stat().st_size == 0
    with PREDICTIONS_CSV_PATH.open("a", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=PREDICTION_FIELDS)
        if new_file:
            writer.writeheader()
        writer.writerow(row)


def append_xlsx(row: dict[str, object]) -> bool:
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        return False
    if XLSX_PATH.exists():
        workbook = load_workbook(XLSX_PATH)
        sheet = workbook["Responses"] if "Responses" in workbook.sheetnames else workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Responses"
        sheet.append(FIELDS)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:I1"
    sheet.append([row[field] for field in FIELDS])
    workbook.save(XLSX_PATH)
    return True


@app.get("/")
def index() -> Response:
    return Response(INDEX_HTML, mimetype="text/html")


@app.get("/api/config")
def config() -> Response:
    files = part_files()
    return jsonify({
        "parts": [{"file": name, "name": Path(name).stem} for name in files],
        "assembly": "GearBoxAssembly.obj",
        "exploded": "GearBoxAssembly_Exploded.obj",
        "excel_enabled": excel_available(),
        "resolver": {
            "enabled": RESOLVER is not None,
            "name": RESOLVER.name if RESOLVER else "none",
            "model": RESOLVER.model_name if RESOLVER else "",
        },
    })


def excel_available() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


@app.get("/assets/<path:filename>")
def assets(filename: str) -> Response:
    return send_from_directory(ASSET_DIR, filename)


@app.get("/parts/<path:filename>")
def parts(filename: str) -> Response:
    if filename not in part_files():
        abort(404)
    return send_from_directory(PART_DIR, filename)


@app.post("/api/responses")
def save_response() -> Response:
    data = request.get_json(silent=True) or {}
    filename = str(data.get("part_file", ""))
    action = str(data.get("action", "response"))
    if filename not in part_files():
        return jsonify({"error": "Unknown part file"}), 400
    if action not in {"response", "skip"}:
        return jsonify({"error": "Invalid action"}), 400
    description = clean_cell(data.get("description"))
    if action == "response" and not description:
        return jsonify({"error": "Description is required"}), 400
    try:
        presentation = max(1, int(data.get("presentation_number", 1)))
        target = max(1, min(100, int(data.get("target_presentations", 3))))
    except (TypeError, ValueError):
        return jsonify({"error": "Invalid presentation count"}), 400
    row = {
        "response_id": str(uuid.uuid4()),
        "timestamp_utc": datetime.now(timezone.utc).isoformat(),
        "participant": clean_cell(data.get("participant"), 200),
        "part_file": filename,
        "part_name": Path(filename).stem,
        "description": description,
        "presentation_number": presentation,
        "target_presentations": target,
        "action": action,
    }
    prediction_row = None
    prediction = data.get("prediction")
    if action == "response" and isinstance(prediction, dict):
        if RESOLVER is None:
            return jsonify({"error": "Cannot record a prediction while the resolver is disabled"}), 400
        predicted_file = str(prediction.get("predicted_part_file") or "")
        if predicted_file and predicted_file not in part_files():
            return jsonify({"error": "Unknown predicted part file"}), 400
        try:
            confidence = optional_finite_number(prediction.get("confidence"))
            mapping_score = optional_finite_number(prediction.get("mapping_score"))
            latency_ms = optional_finite_number(prediction.get("latency_ms"))
        except (TypeError, ValueError) as exc:
            return jsonify({"error": str(exc)}), 400
        prediction_row = {
            "response_id": row["response_id"],
            "timestamp_utc": row["timestamp_utc"],
            "participant": row["participant"],
            "expression": description,
            "ground_truth_part_file": filename,
            "predicted_part_file": clean_cell(predicted_file, 500),
            "predicted_mesh_name": clean_cell(prediction.get("predicted_mesh_name"), 500),
            "is_correct": bool(predicted_file) and predicted_file == filename,
            "confidence": confidence,
            "mapping_score": mapping_score,
            "latency_ms": latency_ms,
            "resolver": RESOLVER.name if RESOLVER else clean_cell(prediction.get("resolver"), 100),
            "model": RESOLVER.model_name if RESOLVER else clean_cell(prediction.get("model"), 500),
            "raw_output": clean_cell(prediction.get("raw_output"), 5000),
        }
    with WRITE_LOCK:
        append_csv(row)
        wrote_xlsx = append_xlsx(row)
        if prediction_row:
            append_prediction_csv(prediction_row)
    return jsonify({"ok": True, "response_id": row["response_id"], "xlsx": wrote_xlsx})


@app.post("/api/resolve")
def resolve_expression() -> Response:
    """Ground an expression without exposing the sampled target to the model."""
    if RESOLVER is None:
        return jsonify({"error": "No referring-expression resolver is enabled"}), 503
    data = request.get_json(silent=True) or {}
    expression = str(data.get("expression") or "").strip()[:2000]
    if not expression:
        return jsonify({"error": "Expression is required"}), 400
    image_url = data.get("image")
    if not isinstance(image_url, str) or not image_url.startswith("data:image/"):
        return jsonify({"error": "A data-URL scene image is required"}), 400
    try:
        encoded = image_url.split(",", 1)[1]
        image_bytes = base64.b64decode(encoded, validate=True)
    except (IndexError, ValueError, binascii.Error):
        return jsonify({"error": "Invalid scene image"}), 400
    if not image_bytes or len(image_bytes) > 12 * 1024 * 1024:
        return jsonify({"error": "Scene image must be between 1 byte and 12 MB"}), 400

    values = data.get("candidates")
    if not isinstance(values, list) or not 1 <= len(values) <= 100:
        return jsonify({"error": "Between 1 and 100 visible candidates are required"}), 400
    try:
        candidates = [Candidate.from_payload(value) for value in values]
    except (TypeError, ValueError) as exc:
        return jsonify({"error": str(exc)}), 400
    known_parts = set(part_files())
    if any(not item.part_file or item.part_file not in known_parts for item in candidates):
        return jsonify({"error": "Candidates contain an unknown part file"}), 400

    try:
        result = RESOLVER.resolve(
            image_bytes=image_bytes,
            expression=expression,
            candidates=candidates,
        )
    except Exception as exc:
        app.logger.exception("Referring-expression inference failed")
        return jsonify({"error": f"Resolver failed: {exc}"}), 500
    return jsonify(result.to_dict())


INDEX_HTML = r'''<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Gearbox Part Description Study</title>
  <script src="https://cdn.babylonjs.com/babylon.js"></script>
  <script src="https://cdn.babylonjs.com/loaders/babylonjs.loaders.min.js"></script>
  <style>
    :root { color-scheme: dark; --page:#101217; --panel:#171a20; --line:#343a46; --text:#f4f6fa; --field:#0e1117; --muted:#cbd2de; }
    body.light { color-scheme:light; --page:#fff; --panel:#f5f6f8; --line:#cbd0d8; --text:#17202d; --field:#fff; --muted:#485365; }
    * { box-sizing: border-box; }
    html, body { margin:0; width:100%; height:100%; overflow:hidden; font:14px/1.35 system-ui,sans-serif; background:var(--page); color:var(--text); }
    #app { display:grid; grid-template-columns:1fr 1fr; height:100%; }
    .viewer { position:relative; min-width:0; border-right:1px solid var(--line); }
    canvas { width:100%; height:100%; display:block; outline:none; touch-action:none; }
    .bar { position:absolute; z-index:2; top:12px; left:12px; right:12px; display:flex; gap:8px; align-items:center; flex-wrap:wrap; padding:9px; border:1px solid #ffffff24; border-radius:10px; background:#11151ddd; backdrop-filter:blur(8px); }
    .bar strong { margin-right:auto; }
    button, input, textarea { font:inherit; }
    button { color:var(--text); border:1px solid var(--line); border-radius:7px; background:var(--panel); padding:7px 11px; cursor:pointer; }
    button:hover { background:#384151; } button.primary { background:#2869db; } button.primary:hover { background:#3479ee; }
    button:disabled { opacity:.45; cursor:not-allowed; }
    #right { display:grid; grid-template-rows:minmax(0,1fr) auto; min-width:0; }
    #partView { position:relative; min-height:240px; }
    #form { padding:14px 16px 16px; background:var(--panel); border-top:1px solid var(--line); }
    .formrow { display:grid; grid-template-columns:minmax(160px,1fr) 150px; gap:12px; margin-bottom:10px; }
    label { display:block; color:var(--muted); font-size:12px; }
    input, textarea { width:100%; margin-top:4px; color:var(--text); background:var(--field); border:1px solid var(--line); border-radius:7px; padding:9px; }
    textarea { height:74px; resize:vertical; }
    .actions { display:flex; gap:8px; align-items:center; margin-top:9px; }
    #prediction { display:none; margin-top:10px; padding:9px 11px; border:1px solid var(--line); border-radius:7px; background:var(--field); }
    #prediction.correct { display:block; border-color:#35a65b; }
    #prediction.incorrect { display:block; border-color:#d46a6a; }
    #prediction.neutral { display:block; border-color:#4c91d9; }
    #status { margin-left:auto; color:#abb5c5; min-height:20px; }
    #partLabel { color:#c9d8ff; overflow-wrap:anywhere; }
    #progress { color:#aeb8c8; }
    .hint { position:absolute; bottom:10px; left:12px; z-index:2; color:#d4dae4; background:#11151dcc; padding:6px 9px; border-radius:7px; pointer-events:none; }
    @media (max-width:850px) { #app { grid-template-columns:1fr; grid-template-rows:1fr 1fr; } #right { min-height:0; } .viewer { border-right:0; border-bottom:1px solid var(--line); } }
  </style>
</head>
<body>
<div id="app">
  <section class="viewer">
    <canvas id="assemblyCanvas"></canvas>
    <div class="bar"><strong>Full assembly</strong><button id="assembledBtn" class="primary">Assembled</button><button id="explodedBtn">Exploded</button><button id="themeBtn">Light theme</button><button id="resetAssembly">Reset view</button></div>
    <div class="hint">The sampled part is boxed here. Click any component to inspect it.</div>
  </section>
  <section id="right">
    <div id="partView">
      <canvas id="partCanvas"></canvas>
      <div class="bar"><strong id="partLabel">Loading parts…</strong><span id="progress"></span><button id="resetPart">Reset view</button></div>
    </div>
    <div id="form">
      <div class="formrow">
        <label>Participant name<input id="participant" autocomplete="name" placeholder="Enter name"></label>
        <label>Descriptions per part<input id="target" type="number" min="1" max="100" value="3"></label>
      </div>
      <label>Describe the highlighted part<textarea id="description" placeholder="Type the participant's description…"></textarea></label>
      <div id="prediction"></div>
      <div class="actions"><button id="evaluate">Ask model</button><button id="submit" class="primary">Save &amp; continue</button><button id="skip">Skip</button><button id="newRound">Restart sampling</button><span id="status"></span></div>
    </div>
  </section>
</div>
<script>
(() => {
  const $ = id => document.getElementById(id);
  const assemblyEngine = new BABYLON.Engine($('assemblyCanvas'), true, {preserveDrawingBuffer:true, stencil:true});
  const partEngine = new BABYLON.Engine($('partCanvas'), true, {preserveDrawingBuffer:true, stencil:true});
  let assemblyScene, assemblyCamera, partScene, currentPart, selectedAssemblyObject=null, parts=[], bag=[], counts={}, skipped={}, presentation=1;
  let resolverConfig={enabled:false,name:'none',model:''}, currentPrediction=null;
  let activeAssemblyFile='GearBoxAssembly.obj'; const assemblyModels={}, assemblyLoads={};
  const colors = {
    black:[.02,.02,.025], white:[.93,.93,.93], red:[.82,.035,.035], green:[.03,.55,.12],
    blue:[.035,.18,.86], brown:[.533,.322,.153], silver:[.53,.56,.60]
  };
  const assemblyParts = {
    'BaseBoard.stl':['Part_1.013'],
    'Bearing.stl':['Part_1','Part_1.005','Part_1.008','Part_1.016','Part_1.018','Part_1.019','Part_1.023','Part_1.024'],
    'Handle.stl':['Part_1.022'],
    'Row1_GearRod.stl':['Part_1.027'], 'Row1_GearStand_Left.stl':['Part_1.004'], 'Row1_GearStand_Right.stl':['Part_1.002'],
    'Row1_Gear_Left.stl':['Spur_gear_(42_teeth).001'], 'Row1_Screws.stl':['90751A122_18-8_Stainless_Steel_Socket_Head_Screws','90751A122_18-8_Stainless_Steel_Socket_Head_Screws.001'],
    'Row2_GearRod.stl':['Part_1.029'], 'Row2_GearStand_Left.stl':['Part_1.026'], 'Row2_GearStand_Right.stl':['Part_1.014'],
    'Row2_Gear_Left.stl':['Spur_gear_(20_teeth).001'], 'Row2_Gear_Right.stl':['Spur_gear_(32_teeth)'], 'Row2_Screws.stl':['90358A118_Ultra-Low-Profile_Socket_Head_Screw','90358A118_Ultra-Low-Profile_Socket_Head_Screw.001'],
    'Row3_GearRod.stl':['Part_1.028'], 'Row3_GearStand_Left.stl':['Part_1.015'], 'Row3_GearStand_Right.stl':['Part_1.003'],
    'Row3_Gear_Left.stl':['Spur_gear_(20_teeth)'], 'Row3_Gear_Right.stl':['Spur_gear_(24_teeth)'], 'Row3_Screws.stl':['90942A121_High-Strength_A286_Stainless_Steel_Button_Head_Torx_Screw','90942A121_High-Strength_A286_Stainless_Steel_Button_Head_Torx_Screw.001'],
    'Row4_Gear_Left.stl':['Spur_gear_(42_teeth)'], 'Row4_GearRod.stl':['Part_1.021'],
    'Row4_GearStand_Left.stl':['Part_1.007'], 'Row4_GearStand_Right.stl':['Part_1.009'],
    'Row4_Screws.stl':['92000A107_Passivated_18-8_Stainless_Steel_Pan_Head_Phillips_Screws','92000A107_Passivated_18-8_Stainless_Steel_Pan_Head_Phillips_Screws.001'],
    'WoodenPin.stl':['Part_1.001','Part_1.006','Part_1.010','Part_1.011','Part_1.012','Part_1.017','Part_1.020']
  };
  function sceneColor(){return document.body.classList.contains('light') ? new BABYLON.Color4(1,1,1,1) : new BABYLON.Color4(.055,.065,.085,1);}
  function setupScene(engine, canvas) {
    const scene = new BABYLON.Scene(engine); scene.clearColor = sceneColor();
    scene.ambientColor = new BABYLON.Color3(.1,.1,.1);
    scene.imageProcessingConfiguration.exposure = 1.0;
    scene.imageProcessingConfiguration.contrast = 1.12;
    const camera = new BABYLON.ArcRotateCamera('camera', -Math.PI/2, Math.PI/2.6, 4, BABYLON.Vector3.Zero(), scene);
    camera.attachControl(canvas, true); camera.wheelPrecision=45; camera.panningSensibility=1200; camera.lowerRadiusLimit=.01;
    const upper=new BABYLON.HemisphericLight('upperFill',new BABYLON.Vector3(0,1,0),scene); upper.intensity=.62; upper.groundColor=new BABYLON.Color3(.18,.2,.24);
    const lower=new BABYLON.HemisphericLight('lowerFill',new BABYLON.Vector3(0,-1,0),scene); lower.intensity=.08; lower.groundColor=new BABYLON.Color3(.1,.11,.13);
    const light = new BABYLON.DirectionalLight('key', new BABYLON.Vector3(-1,-2,-1), scene); light.intensity=.58;
    const cameraLight=new BABYLON.PointLight('cameraFill',BABYLON.Vector3.Zero(),scene); cameraLight.parent=camera; cameraLight.intensity=.24;
    return {scene,camera};
  }
  function frame(scene, meshes, camera=scene.activeCamera) {
    const visible=meshes.filter(m=>m.getTotalVertices && m.getTotalVertices()>0 && m.isEnabled());
    if(!visible.length) return;
    let min=new BABYLON.Vector3(Infinity,Infinity,Infinity), max=new BABYLON.Vector3(-Infinity,-Infinity,-Infinity);
    visible.forEach(m=>{ m.computeWorldMatrix(true); const b=m.getBoundingInfo().boundingBox; min=BABYLON.Vector3.Minimize(min,b.minimumWorld); max=BABYLON.Vector3.Maximize(max,b.maximumWorld); });
    const center=min.add(max).scale(.5), size=max.subtract(min).length(); camera.setTarget(center); camera.radius=Math.max(size*.72,.02); camera.minZ=Math.max(size/10000,.0001); camera.maxZ=Math.max(size*20,100);
  }
  function showBoxes(meshes, scene, kind='target') {
    scene.meshes.forEach(m=>{if(m.showBoundingBox)m.showBoundingBox=false;});
    scene.forceShowBoundingBoxes=false;
    meshes.slice(0,1).forEach(mesh=>mesh.showBoundingBox=true);
    const renderer=scene.getBoundingBoxRenderer();
    const light=document.body.classList.contains('light');
    renderer.frontColor=kind==='prediction' ? new BABYLON.Color3(.1,.85,1) : (light ? new BABYLON.Color3(.02,.12,.65) : new BABYLON.Color3(1,.82,.05));
    renderer.backColor=renderer.frontColor;
  }
  function highlightSampledPart() {
    if(!assemblyScene||!currentPart)return;
    const model=assemblyModels[activeAssemblyFile];
    const mapped=assemblyParts[currentPart.file]||[];
    const wanted=selectedAssemblyObject&&mapped.includes(selectedAssemblyObject) ? selectedAssemblyObject : mapped[0];
    showBoxes(model ? model.meshes.filter(m=>m.name===wanted).slice(0,1) : [],assemblyScene);
  }
  function highlightPrediction() {
    if(!assemblyScene||!currentPrediction)return;
    const model=assemblyModels[activeAssemblyFile];
    const wanted=currentPrediction.predicted_mesh_name;
    showBoxes(model&&wanted ? model.meshes.filter(m=>m.name===wanted).slice(0,1) : [],assemblyScene,'prediction');
  }
  function highlightStudySelection() {
    if(currentPrediction)highlightPrediction(); else highlightSampledPart();
  }
  function clearPrediction() {
    currentPrediction=null;
    $('prediction').className=''; $('prediction').textContent='';
    if(assemblyScene&&currentPart)highlightSampledPart();
  }
  function partForAssemblyObject(name) {
    const entry=Object.entries(assemblyParts).find(([,names])=>names.includes(name));
    return entry ? parts.find(part=>part.file===entry[0]) : null;
  }
  function materialKey(file) {
    if(file==='BaseBoard.stl'||file==='Bearing.stl'||file==='Row2_Screws.stl'||file==='Row3_Screws.stl') return 'black';
    if(file==='Handle.stl'||file.startsWith('Row1_')&&!file.includes('Screws')) return 'white';
    if(file==='Row1_Screws.stl'||file==='Row4_Screws.stl') return 'silver';
    if(file==='WoodenPin.stl') return 'brown'; if(file.startsWith('Row2')) return 'red';
    if(file.startsWith('Row3')) return 'green'; if(file.startsWith('Row4')) return 'blue'; return 'white';
  }
  function applyPartMaterial(scene, meshes, file) {
    const key=materialKey(file), c=colors[key], mat=new BABYLON.PBRMaterial('part-'+key,scene);
    mat.albedoColor=new BABYLON.Color3(...c); mat.metallic=key==='silver' ? .9 : 0; mat.roughness=key==='silver' ? .28 : .55;
    meshes.forEach(m=>{if(m.getTotalVertices()>0)m.material=mat;});
  }
  async function loadAssembly(file, activate=true) {
    if(!assemblyScene) {
      const made=setupScene(assemblyEngine,$('assemblyCanvas')); assemblyScene=made.scene; assemblyCamera=made.camera;
      assemblyScene.onPointerObservable.add(info=>{
        if(info.type!==BABYLON.PointerEventTypes.POINTERPICK) return;
        const pick=info.pickInfo;
        if(pick && pick.hit && pick.pickedMesh && pick.pickedMesh.getTotalVertices()>0) {
          const part=partForAssemblyObject(pick.pickedMesh.name);
          if(part) {
            bag=bag.filter(item=>item.file!==part.file);
            loadPart(part,pick.pickedMesh.name).catch(showError);
          } else showError('This assembly object has no matching study part.');
        }
      });
    }
    if(activate) {
      activeAssemblyFile=file;
      Object.values(assemblyModels).forEach(model=>model.root.setEnabled(false));
    }
    if(!assemblyModels[file]) {
      if(activate)$('status').textContent='Loading '+(file.includes('Exploded')?'exploded':'assembled')+' model once…';
      if(!assemblyLoads[file]) assemblyLoads[file]=(async()=>{
        const result=await BABYLON.SceneLoader.ImportMeshAsync('', '/assets/', file, assemblyScene);
        const root=new BABYLON.TransformNode('cached-'+file,assemblyScene);
        result.meshes.forEach(m=>{if(!m.parent)m.parent=root;});
        [...new Set(result.meshes.map(m=>m.material).filter(Boolean))].forEach(mat=>{
          if(mat.specularColor)mat.specularColor=new BABYLON.Color3(.12,.12,.12);
          if('specularPower' in mat)mat.specularPower=64;
        });
        assemblyModels[file]={root,meshes:result.meshes};
        if(activeAssemblyFile!==file)root.setEnabled(false);
      })();
      await assemblyLoads[file];
      if(activate)$('status').textContent='Model cached';
    }
    if(!activate)return;
    const model=assemblyModels[file];
    if(activeAssemblyFile!==file)return;
    model.root.setEnabled(true);
    frame(assemblyScene,model.meshes,assemblyCamera);
    highlightStudySelection();
  }
  async function loadPart(part, assemblyObject=null) {
    clearPrediction();
    currentPart=part; selectedAssemblyObject=assemblyObject; presentation=(counts[part.file]||0)+1;
    $('submit').disabled=false; $('skip').disabled=false;
    highlightStudySelection();
    $('partLabel').textContent=part.name; updateProgress();
    if(partScene) partScene.dispose();
    const made=setupScene(partEngine,$('partCanvas')); partScene=made.scene;
    const result=await BABYLON.SceneLoader.ImportMeshAsync('', '/parts/', encodeURIComponent(part.file), partScene);
    applyPartMaterial(partScene,result.meshes,part.file); frame(partScene,result.meshes,made.camera);
    highlightSampledPart();
    $('description').value=''; $('description').focus();
  }
  function refillBag() {
    const target=Number($('target').value)||3;
    bag=parts.filter(p=>(counts[p.file]||0)<target&&!skipped[p.file]&&(!currentPart||p.file!==currentPart.file));
    for(let i=bag.length-1;i>0;i--){const j=Math.floor(Math.random()*(i+1));[bag[i],bag[j]]=[bag[j],bag[i]];}
  }
  function nextPart(forceNew=false) {
    const target=Number($('target').value)||3;
    if(!forceNew&&currentPart&&(counts[currentPart.file]||0)<target&&!skipped[currentPart.file]) { loadPart(currentPart,selectedAssemblyObject).catch(showError); return; }
    if(!bag.length) refillBag();
    if(!bag.length){ currentPart=null; $('partLabel').textContent='Sampling complete'; $('progress').textContent=''; $('submit').disabled=true; $('skip').disabled=true; return; }
    $('submit').disabled=false; $('skip').disabled=false; loadPart(bag.pop()).catch(showError);
  }
  function updateProgress(){const target=Number($('target').value)||3, completed=Object.values(counts).reduce((a,b)=>a+b,0); $('progress').textContent=`showing ${presentation}/${target} · ${completed}/${parts.length*target} saved`;}
  function projectedBox(mesh) {
    mesh.computeWorldMatrix(true);
    const vectors=mesh.getBoundingInfo().boundingBox.vectorsWorld;
    const viewport=assemblyCamera.viewport.toGlobal(assemblyEngine.getRenderWidth(),assemblyEngine.getRenderHeight());
    const points=vectors.map(point=>BABYLON.Vector3.Project(point,BABYLON.Matrix.Identity(),assemblyScene.getTransformMatrix(),viewport));
    let x1=Math.max(0,Math.min(...points.map(p=>p.x))), y1=Math.max(0,Math.min(...points.map(p=>p.y)));
    let x2=Math.min(assemblyEngine.getRenderWidth(),Math.max(...points.map(p=>p.x))), y2=Math.min(assemblyEngine.getRenderHeight(),Math.max(...points.map(p=>p.y)));
    if(![x1,y1,x2,y2].every(Number.isFinite)||x2-x1<2||y2-y1<2)return null;
    return [x1,y1,x2,y2];
  }
  function visibleCandidates() {
    const model=assemblyModels[activeAssemblyFile]; if(!model)return [];
    const result=[];
    Object.entries(assemblyParts).forEach(([partFile,names])=>names.forEach(meshName=>{
      model.meshes.filter(mesh=>mesh.name===meshName&&mesh.isEnabled()&&mesh.getTotalVertices()>0).forEach(mesh=>{
        const bbox=projectedBox(mesh); if(bbox)result.push({part_file:partFile,mesh_name:meshName,bbox});
      });
    }));
    return result;
  }
  async function evaluateExpression() {
    if(!currentPart||!resolverConfig.enabled)return;
    const expression=$('description').value.trim();
    if(!expression){showError('Enter a description before asking the model.');return;}
    const evaluatedPartFile=currentPart.file;
    setBusy(true); $('status').textContent='Model is resolving the expression…';
    try {
      showBoxes([],assemblyScene); assemblyScene.render();
      const candidates=visibleCandidates();
      if(!candidates.length)throw new Error('No visible assembly candidates could be projected.');
      const image=$('assemblyCanvas').toDataURL('image/jpeg',.9);
      const res=await fetch('/api/resolve',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify({expression,image,candidates})});
      const data=await res.json(); if(!res.ok)throw new Error(data.error||'Model evaluation failed');
      if(!currentPart||currentPart.file!==evaluatedPartFile)throw new Error('The sampled part changed during evaluation; please ask the model again.');
      currentPrediction=data; const correct=data.predicted_part_file===currentPart.file;
      $('prediction').className=data.predicted_part_file?(correct?'correct':'incorrect'):'neutral';
      const confidence=data.confidence==null?'n/a':Number(data.confidence).toFixed(3);
      const mapping=data.mapping_score==null?'n/a':Number(data.mapping_score).toFixed(3);
      const label=data.predicted_part_file||'No grounded part';
      $('prediction').textContent=`Model: ${label} · ${correct?'correct':'incorrect'} · confidence ${confidence} · box/mesh match ${mapping} · ${Number(data.latency_ms).toFixed(0)} ms`;
      $('status').textContent='Model result ready; save to record it'; highlightPrediction();
    } catch(e){clearPrediction();showError(e);} finally{setBusy(false);}
  }
  async function record(action) {
    if(!currentPart) return;
    const description=$('description').value.trim(); if(action==='response'&&!description){showError('Please enter a description, or use Skip.');return;}
    setBusy(true); const payload={participant:$('participant').value,part_file:currentPart.file,description,presentation_number:presentation,target_presentations:Number($('target').value)||3,action,prediction:currentPrediction};
    try { const res=await fetch('/api/responses',{method:'POST',headers:{'Content-Type':'application/json'},body:JSON.stringify(payload)}); const data=await res.json(); if(!res.ok)throw new Error(data.error||'Save failed'); if(action==='response')counts[currentPart.file]=(counts[currentPart.file]||0)+1; else skipped[currentPart.file]=true; $('status').textContent=action==='skip'?'Part skipped':'Response saved'; nextPart(action==='skip'); }
    catch(e){showError(e);} finally{setBusy(false);}
  }
  function setBusy(v){$('submit').disabled=v;$('skip').disabled=v;$('evaluate').disabled=v||!resolverConfig.enabled;}
  function showError(e){$('status').textContent=e.message||String(e);$('status').style.color='#ff8d8d';setTimeout(()=>{$('status').style.color='';},3500);}
  async function init(){
    const cfg=await fetch('/api/config').then(r=>r.json()); parts=cfg.parts; resolverConfig=cfg.resolver;
    if(!parts.length)throw new Error('No STL files found in completed/colored_stl');
    $('evaluate').disabled=!resolverConfig.enabled;
    $('evaluate').title=resolverConfig.enabled?`${resolverConfig.name}: ${resolverConfig.model}`:'Start the app with --resolver florence2 to enable model evaluation';
    if(!cfg.excel_enabled)$('status').textContent='Saving CSV (install openpyxl for Excel too)';
    await loadAssembly(cfg.assembly); refillBag(); nextPart();
    loadAssembly(cfg.exploded,false).catch(showError);
  }
  $('assembledBtn').onclick=()=>{ $('assembledBtn').classList.add('primary');$('explodedBtn').classList.remove('primary');loadAssembly('GearBoxAssembly.obj').catch(showError); };
  $('explodedBtn').onclick=()=>{ $('explodedBtn').classList.add('primary');$('assembledBtn').classList.remove('primary');loadAssembly('GearBoxAssembly_Exploded.obj').catch(showError); };
  $('resetAssembly').onclick=()=>{const model=assemblyModels[activeAssemblyFile];if(model)frame(assemblyScene,model.meshes);};
  $('resetPart').onclick=()=>partScene&&frame(partScene,partScene.meshes);
  $('themeBtn').onclick=()=>{const light=document.body.classList.toggle('light');$('themeBtn').textContent=light?'Dark theme':'Light theme';if(assemblyScene){assemblyScene.clearColor=sceneColor();highlightStudySelection();}if(partScene)partScene.clearColor=sceneColor();};
  $('evaluate').onclick=evaluateExpression; $('submit').onclick=()=>record('response'); $('skip').onclick=()=>record('skip');
  $('newRound').onclick=()=>{counts={};skipped={};currentPart=null;refillBag();nextPart(true);$('status').textContent='Sampling restarted';};
  $('target').onchange=()=>{ $('target').value=Math.max(1,Math.min(100,Number($('target').value)||3)); refillBag(); updateProgress(); if(!currentPart||(counts[currentPart.file]||0)>=$('target').value)nextPart(true); };
  $('description').addEventListener('keydown',e=>{if((e.ctrlKey||e.metaKey)&&e.key==='Enter')record('response');});
  $('description').addEventListener('input',()=>{if(currentPrediction)clearPrediction();});
  window.addEventListener('resize',()=>{assemblyEngine.resize();partEngine.resize();});
  assemblyEngine.runRenderLoop(()=>{if(assemblyScene)assemblyScene.render();}); partEngine.runRenderLoop(()=>{if(partScene)partScene.render();}); init().catch(showError);
})();
</script>
</body>
</html>'''


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=5050)
    parser.add_argument(
        "--resolver",
        choices=("none", "florence2"),
        default="none",
        help="Ground participant expressions with the selected local backend",
    )
    parser.add_argument(
        "--model",
        default=None,
        help="Optional model ID override for the selected resolver",
    )
    args = parser.parse_args()
    if not PART_DIR.is_dir():
        raise SystemExit(f"Missing colored STL directory: {PART_DIR}")
    RESOLVER = create_resolver(args.resolver, args.model)
    print(f"Open http://{args.host}:{args.port}")
    print(f"Responses: {CSV_PATH}")
    if RESOLVER:
        print(f"Resolver: {RESOLVER.name} ({RESOLVER.model_name}); model loads on first request")
        print(f"Predictions: {PREDICTIONS_CSV_PATH}")
    else:
        print("Resolver: disabled (use --resolver florence2 to enable)")
    if not excel_available():
        print("Tip: install openpyxl to also save an .xlsx workbook")
    app.run(host=args.host, port=args.port, debug=False, threaded=True)
